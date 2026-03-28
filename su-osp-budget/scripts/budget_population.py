# Copyright 2026 Duncan Brown
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
OSP Budget Template Population Script

This script safely populates the Syracuse University OSP Budget Template
with data while preserving all formulas and formatting.

IMPORTANT: This script should only be used by Claude Opus 4.5 models.
Complex budget calculations and natural language interpretation require
the most capable model to avoid errors.
"""

from openpyxl import load_workbook
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any
from enum import Enum
import shutil
import os
import zipfile
import tempfile


def preserve_external_links(template_path: str, output_path: str) -> int:
    """
    Replace external link files in output with originals from template.
    
    openpyxl corrupts external link files by stripping namespaces and elements
    (e.g., xmlns:xxl21, alternateUrls) that Excel requires. This causes Excel 
    to report "We found a problem with some content" errors.
    
    This function restores the original external link files from the template
    after openpyxl has saved the workbook, preserving the data changes while
    fixing the corrupted external links.
    
    Args:
        template_path: Path to the original template file
        output_path: Path to the openpyxl-modified file to fix
        
    Returns:
        Number of files restored
    """
    temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(temp_fd)
    
    # Get original external link files from template
    original_ext_files = {}
    with zipfile.ZipFile(template_path, 'r') as ztemplate:
        for item in ztemplate.infolist():
            if 'externalLinks/' in item.filename:
                original_ext_files[item.filename] = ztemplate.read(item.filename)
    
    restored_count = 0
    try:
        with zipfile.ZipFile(output_path, 'r') as zin:
            with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if 'externalLinks/' in item.filename:
                        # Use original from template if available
                        if item.filename in original_ext_files:
                            zout.writestr(item, original_ext_files[item.filename])
                            restored_count += 1
                        else:
                            # Keep file from output if not in template
                            zout.writestr(item, zin.read(item.filename))
                    else:
                        # Keep all non-external-link files from output
                        zout.writestr(item, zin.read(item.filename))
                
                # Add any template external link files that aren't in output
                output_files = set(zin.namelist())
                for filename, content in original_ext_files.items():
                    if filename not in output_files:
                        zout.writestr(filename, content)
                        restored_count += 1
        
        shutil.move(temp_path, output_path)
        return restored_count
        
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise e

# === DROPDOWN VALUES (must be exact matches) ===

class SponsorType(Enum):
    FEDERAL_NIH = "Federal - NIH"
    FEDERAL_OTHER = "Federal - Other"
    NON_FEDERAL = "Fnd/Prof Soc"

class Designation(Enum):
    TENURE_TRACK_ACAD_SUM = "Tenure/Track Faculty - Acad/Sum"
    TENURE_TRACK_CALENDAR = "Tenure/Track Faculty - Calendar"
    NON_TENURE_FULL_TIME = "Non-Tenure Track Faculty - Full Time"
    NON_TENURE_HOURLY = "Non-Tenure Track Faculty - Hourly"
    SR_RESEARCH_ASSOCIATE = "Senior Research Associate (Staff)"
    QUALIFIED_STAFF = "Qualified Staff Non-Ph.D."

class SeniorRole(Enum):
    PI = "PI"
    CO_PI = "Co-PI"
    CO_INVESTIGATOR = "Co-Investigator"
    SENIOR_KEY = "Senior/Key Personnel"

class OtherRole(Enum):
    POSTDOC = "Postdoctoral Associates"
    RA_FULL_TIME = "Research Associate - Full Time"
    RA_HOURLY = "Research Associate - Hourly"
    OTHER_PROF_STAFF = "Other Professional Staff"
    GRAD_ASSISTANT = "Graduate Assistants"
    HOURLY_STUDENT = "Hourly Students (Grad/Undergrad)"
    SECRETARIAL = "Secretarial/Clerical"
    OTHER_TEMP = "Other (Temp, wages)"
    EXTRA_SERVICE = "Extra Service & Overload"
    ADJUNCT = "Adjunct Faculty"

class FAType(Enum):
    MTDC_FED = "MTDC-Fed"
    MTDC_NONFED = "MTDC-NonFed"
    TDC = "TDC"
    SWFB = "SWFB"

class FARate(Enum):
    RESEARCH_ON = "Research - On Campus - 49.50%"
    INSTRUCTION_ON = "Instruction - On Campus - 34.00%"
    OTHER_ON = "Other - On Campus - 35.00%"
    RESEARCH_OFF = "Research - Off Campus - 26.00%"
    INSTRUCTION_OFF = "Instruction - Off Campus - 26.00%"
    OTHER_OFF = "Other - Off Campus - 26.00%"


# === DATA CLASSES ===

@dataclass
class ProjectInfo:
    start_date: str
    duration_years: int  # 1-5
    sponsor_type: SponsorType
    
    def validate(self) -> List[str]:
        errors = []
        if self.duration_years not in [1, 2, 3, 4, 5]:
            errors.append(f"Duration must be 1-5, got {self.duration_years}")
        return errors


@dataclass
class SeniorPerson:
    first_name: str
    last_name: str
    role: SeniorRole
    designation: Designation
    base_salary: float
    cal_months: float = 0
    acad_months: float = 0
    sum_months: float = 0
    prefix: str = ""
    middle: str = ""
    suffix: str = ""
    justification: str = ""
    
    # Fields to help generate justification automatically
    course_buyouts: int = 0              # Number of courses bought out
    po_exception: bool = False           # Program officer approved exception to 2-month rule
    po_exception_note: str = ""          # Details of exception
    
    def generate_justification(self) -> str:
        """Generate automatic justification text based on effort details."""
        parts = []
        
        total_months = self.cal_months + self.acad_months + self.sum_months
        
        # Course buyout explanation
        if self.course_buyouts > 0:
            buyout_months = self.course_buyouts * 1.275
            parts.append(f"{self.course_buyouts} course buyout{'s' if self.course_buyouts > 1 else ''} "
                        f"at 15% of academic year = {buyout_months:.3f} academic months")
        
        # Explain the month breakdown
        month_parts = []
        if self.acad_months > 0 and self.course_buyouts == 0:
            month_parts.append(f"{self.acad_months} academic months")
        if self.sum_months > 0:
            month_parts.append(f"{self.sum_months} summer months")
        if self.cal_months > 0:
            month_parts.append(f"{self.cal_months} calendar months")
        
        if month_parts and self.course_buyouts > 0:
            parts.append(f"Plus {', '.join(month_parts)}")
        elif month_parts and self.course_buyouts == 0:
            parts.append(f"Effort: {', '.join(month_parts)}")
        
        # Total months check
        if total_months > 0:
            parts.append(f"Total: {total_months:.3f} months")
        
        # PO exception note
        if self.po_exception:
            if self.po_exception_note:
                parts.append(f"PO exception: {self.po_exception_note}")
            else:
                parts.append("Program officer approved exception to 2-month rule")
        
        # Use manual justification if provided, otherwise generate
        if self.justification:
            return self.justification
        elif parts:
            return ". ".join(parts) + "."
        else:
            return ""
    
    def validate(self) -> List[str]:
        errors = []
        if not self.first_name:
            errors.append("First name required")
        if not self.last_name:
            errors.append("Last name required")
        if self.cal_months < 0 or self.cal_months > 12:
            errors.append(f"Cal months must be 0-12, got {self.cal_months}")
        if self.acad_months < 0 or self.acad_months > 8.5:
            errors.append(f"Acad months must be 0-8.5, got {self.acad_months}")
        if self.sum_months < 0 or self.sum_months > 3.5:
            errors.append(f"Sum months must be 0-3.5, got {self.sum_months}")
        return errors


@dataclass
class OtherPerson:
    first_name: str
    last_name: str
    role: OtherRole
    requested_salary: float
    cal_months: float = 0
    acad_months: float = 0
    sum_months: float = 0
    prefix: str = ""
    middle: str = ""
    suffix: str = ""
    justification: str = ""
    
    # Fields to help generate justification automatically
    # For hourly workers
    hourly_rate: float = 0
    hours_per_week: float = 0
    weeks: float = 0
    
    # For grad students
    annual_stipend: float = 0       # e.g., 28750
    stipend_months_base: float = 9  # GA = 9 months
    
    # For salaried (postdocs, research associates)
    annual_salary: float = 0
    fte: float = 0                  # e.g., 0.5 for half-time
    
    def generate_justification(self) -> str:
        """Generate automatic justification text based on salary calculation."""
        parts = []
        
        if self.role == OtherRole.HOURLY_STUDENT:
            # Hourly calculation
            if self.hourly_rate > 0 and self.hours_per_week > 0 and self.weeks > 0:
                total_hours = self.hours_per_week * self.weeks
                calc_salary = self.hourly_rate * total_hours
                parts.append(f"${self.hourly_rate:.2f}/hr × {self.hours_per_week} hrs/wk × {self.weeks} weeks "
                           f"= {total_hours:.0f} hours = ${calc_salary:,.0f}")
        
        elif self.role == OtherRole.GRAD_ASSISTANT:
            # Grad student calculation
            if self.annual_stipend > 0:
                monthly_rate = self.annual_stipend / self.stipend_months_base
                
                calc_parts = []
                total_salary = 0
                
                if self.acad_months > 0:
                    acad_salary = self.acad_months * monthly_rate
                    total_salary += acad_salary
                    if self.acad_months == self.stipend_months_base:
                        calc_parts.append(f"Full AY stipend ${self.annual_stipend:,.0f}")
                    else:
                        calc_parts.append(f"{self.acad_months} acad mo × ${monthly_rate:,.2f}/mo = ${acad_salary:,.0f}")
                
                if self.sum_months > 0:
                    sum_salary = self.sum_months * monthly_rate
                    total_salary += sum_salary
                    calc_parts.append(f"{self.sum_months} summer mo × ${monthly_rate:,.2f}/mo = ${sum_salary:,.0f}")
                
                if calc_parts:
                    parts.append(" + ".join(calc_parts))
                    if len(calc_parts) > 1:
                        parts.append(f"Total: ${total_salary:,.0f}")
        
        elif self.role in [OtherRole.POSTDOC, OtherRole.RA_FULL_TIME, OtherRole.OTHER_PROF_STAFF]:
            # Salaried calculation
            if self.annual_salary > 0 and self.fte > 0:
                calc_salary = self.annual_salary * self.fte
                if self.fte == 1.0:
                    parts.append(f"1.0 FTE × ${self.annual_salary:,.0f} annual = ${calc_salary:,.0f}")
                else:
                    parts.append(f"{self.fte} FTE × ${self.annual_salary:,.0f} annual = ${calc_salary:,.0f}")
            elif self.annual_salary > 0 and self.cal_months > 0:
                monthly = self.annual_salary / 12
                calc_salary = monthly * self.cal_months
                parts.append(f"{self.cal_months} months × ${monthly:,.2f}/mo = ${calc_salary:,.0f}")
        
        # Use manual justification if provided, otherwise generate
        if self.justification:
            return self.justification
        elif parts:
            return ". ".join(parts) + "."
        else:
            return ""
    
    def validate(self) -> List[str]:
        errors = []
        if not self.first_name:
            errors.append("First name required")
        if not self.last_name:
            errors.append("Last name required")
        if self.requested_salary < 0:
            errors.append(f"Salary must be positive, got {self.requested_salary}")
        return errors


@dataclass
class PersonnelChange:
    """
    Track a change to personnel in out-years.
    
    Captures INTENT (not just cell values) so we can:
    - Explain changes to the user
    - Detect conflicts
    - Regenerate from state
    - Handle non-linear budget building
    """
    description: str          # Human-readable: "TBD Graduate Student graduates after Year 1"
    person_name: str          # "Jane Smith" or "TBD"
    row: int                  # Row in the sheet (7-14, 20-34 for senior; 53-72 for other)
    section: str              # "senior" or "other"
    change_type: str          # "remove", "modify", "add"
    effective_year: int       # Year change takes effect (2-5)
    end_year: Optional[int] = None   # When change ends (None = through end of project)
    new_values: Dict[str, Any] = field(default_factory=dict)  # {column_letter: value}
    justification: str = ""   # Explanation for column P
    
    def validate(self) -> List[str]:
        errors = []
        if self.section not in ["senior", "other"]:
            errors.append(f"Section must be 'senior' or 'other', got {self.section}")
        if self.change_type not in ["remove", "modify", "add"]:
            errors.append(f"Change type must be 'remove', 'modify', or 'add', got {self.change_type}")
        if self.effective_year < 2 or self.effective_year > 5:
            errors.append(f"Effective year must be 2-5, got {self.effective_year}")
        if self.end_year is not None and self.end_year < self.effective_year:
            errors.append(f"End year ({self.end_year}) cannot be before effective year ({self.effective_year})")
        if self.change_type == "add" and not self.new_values:
            errors.append("Add changes must include new_values")
        return errors
    
    def applies_to_year(self, year: int) -> bool:
        """Check if this change applies to a given budget year."""
        if year < self.effective_year:
            return False
        if self.end_year is not None and year > self.end_year:
            return False
        return True
    
    def get_cells_for_year(self, year: int) -> Dict[str, Any]:
        """
        Get the cell values to write for a given year.
        
        For 'remove': returns cells to zero out
        For 'modify'/'add': returns the new_values
        """
        if not self.applies_to_year(year):
            return {}
        
        if self.change_type == "remove":
            if self.section == "senior":
                # Zero out effort months - salary calculated by formula
                return {"J": 0, "K": 0, "L": 0}
            else:
                # Zero out months AND salary (Other Personnel salary is manual)
                return {"J": 0, "K": 0, "L": 0, "M": 0}
        else:
            return self.new_values.copy()


@dataclass
class BudgetState:
    """
    Complete state of a budget that can be serialized/restored.
    
    This captures the FULL budget intent, allowing:
    - Non-linear budget building (go back and change things)
    - Regeneration of Excel from state
    - Parsing uploaded Excel back to state
    - Validation of conflicts
    """
    project_info: ProjectInfo
    year1_senior: List[SeniorPerson] = field(default_factory=list)
    year1_other: List[OtherPerson] = field(default_factory=list)
    grad_fringe_type: Optional[SponsorType] = None
    personnel_changes: List[PersonnelChange] = field(default_factory=list)
    # TODO: Add non-personnel costs
    
    def add_change(self, change: PersonnelChange) -> List[str]:
        """Add a personnel change, checking for conflicts."""
        errors = change.validate()
        if errors:
            return errors
        
        # Check for conflicts with existing changes
        for existing in self.personnel_changes:
            if existing.row == change.row and existing.section == change.section:
                # Same row - check year overlap
                for year in range(change.effective_year, (change.end_year or 5) + 1):
                    if existing.applies_to_year(year):
                        # Conflict! But might be intentional (e.g., undergrad Y2, grad Y3)
                        # Just warn, don't error
                        pass
        
        self.personnel_changes.append(change)
        return []
    
    def get_changes_for_year(self, year: int) -> List[PersonnelChange]:
        """Get all changes that apply to a specific year."""
        return [c for c in self.personnel_changes if c.applies_to_year(year)]
    
    def summarize_changes(self) -> str:
        """Generate a human-readable summary of all out-year changes."""
        if not self.personnel_changes:
            return "No out-year changes."
        
        lines = ["## Out-Year Personnel Changes\n"]
        for year in range(2, 6):
            year_changes = self.get_changes_for_year(year)
            if year_changes:
                lines.append(f"\n### Year {year}\n")
                for c in year_changes:
                    lines.append(f"- **{c.person_name}** (Row {c.row}): {c.description}")
        
        return "\n".join(lines)
    
    def to_budget_data(self) -> 'BudgetData':
        """Convert to BudgetData for Year 1 population."""
        return BudgetData(
            project_info=self.project_info,
            senior_personnel=self.year1_senior,
            other_personnel=self.year1_other,
            grad_fringe_type=self.grad_fringe_type
        )


@dataclass
class BudgetData:
    """Year 1 budget data (original structure, kept for compatibility)."""
    project_info: ProjectInfo
    senior_personnel: List[SeniorPerson] = field(default_factory=list)
    other_personnel: List[OtherPerson] = field(default_factory=list)
    grad_fringe_type: Optional[SponsorType] = None
    
    def validate(self) -> List[str]:
        errors = []
        errors.extend(self.project_info.validate())
        
        if len(self.senior_personnel) == 0:
            errors.append("At least one senior person (PI) required")
        elif self.senior_personnel[0].role != SeniorRole.PI:
            errors.append("First senior person must be PI")
        
        if len(self.senior_personnel) > 23:  # 8 primary + 15 additional
            errors.append(f"Max 23 senior personnel, got {len(self.senior_personnel)}")
        
        if len(self.other_personnel) > 20:
            errors.append(f"Max 20 other personnel, got {len(self.other_personnel)}")
        
        # Check if grad students exist but no grad fringe type set
        has_grad = any(p.role == OtherRole.GRAD_ASSISTANT for p in self.other_personnel)
        if has_grad and self.grad_fringe_type is None:
            errors.append("Grad fringe type (I43) must be set when Graduate Assistants are included")
        
        for i, person in enumerate(self.senior_personnel):
            for err in person.validate():
                errors.append(f"Senior person {i+1}: {err}")
        
        for i, person in enumerate(self.other_personnel):
            for err in person.validate():
                errors.append(f"Other person {i+1}: {err}")
        
        return errors


# === SAFE CELL DEFINITIONS ===

# These are the ONLY cells that may be written to
SAFE_CELLS_PERSONNEL_YR1 = {
    # Project info
    'I5', 'K5', 'O5',
    # Senior personnel rows 7-14 (data columns only)
    *[f'{col}{row}' for row in range(7, 15) for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'P']],
    # Additional senior personnel rows 20-34
    *[f'{col}{row}' for row in range(20, 35) for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'P']],
    # Grad fringe type
    'I43',
    # Other personnel detail rows 53-72 (including M for salary)
    *[f'{col}{row}' for row in range(53, 73) for col in ['B', 'C', 'D', 'E', 'F', 'G', 'J', 'K', 'L', 'M', 'P']],
}


def is_safe_cell(sheet_name: str, cell: str) -> bool:
    """Check if a cell is safe to write to."""
    if sheet_name == 'Personnel Yr 1':
        return cell in SAFE_CELLS_PERSONNEL_YR1
    # Add other sheets as needed
    return False


# === POPULATION FUNCTIONS ===

def populate_budget(template_path: str, output_path: str, data: BudgetData) -> Dict[str, Any]:
    """
    Populate the OSP Budget Template with the provided data.
    
    Returns a report with:
    - cells_written: list of cells that were written
    - validation_errors: any data validation errors
    - warnings: any warnings (e.g., NIH cap, NSF 2-month rule)
    """
    
    # Validate data first
    validation_errors = data.validate()
    if validation_errors:
        return {
            'success': False,
            'validation_errors': validation_errors,
            'cells_written': [],
            'warnings': []
        }
    
    # Copy template
    shutil.copy(template_path, output_path)
    
    # Load workbook (preserves formulas)
    wb = load_workbook(output_path)
    sheet = wb['Personnel Yr 1']
    
    cells_written = []
    warnings = []
    
    def safe_write(cell: str, value: Any):
        """Write to a cell only if it's in the safe list."""
        if not is_safe_cell('Personnel Yr 1', cell):
            raise ValueError(f"Attempted to write to unsafe cell: {cell}")
        sheet[cell] = value
        cells_written.append((cell, value))
    
    # === Project Info ===
    safe_write('I5', data.project_info.start_date)
    safe_write('K5', data.project_info.duration_years)
    safe_write('O5', data.project_info.sponsor_type.value)
    
    # === Senior Personnel ===
    for i, person in enumerate(data.senior_personnel):
        if i < 8:
            row = 7 + i  # Rows 7-14
        else:
            row = 20 + (i - 8)  # Rows 20-34
        
        if person.prefix:
            safe_write(f'B{row}', person.prefix)
        safe_write(f'C{row}', person.first_name)
        if person.middle:
            safe_write(f'D{row}', person.middle)
        safe_write(f'E{row}', person.last_name)
        if person.suffix:
            safe_write(f'F{row}', person.suffix)
        safe_write(f'G{row}', person.role.value)
        safe_write(f'H{row}', person.designation.value)
        safe_write(f'I{row}', person.base_salary)
        
        if person.cal_months > 0:
            safe_write(f'J{row}', person.cal_months)
        if person.acad_months > 0:
            safe_write(f'K{row}', person.acad_months)
        if person.sum_months > 0:
            safe_write(f'L{row}', person.sum_months)
        
        # Generate and write justification
        justification = person.generate_justification()
        if justification:
            safe_write(f'P{row}', justification)
        
        # Check NSF 2-month rule
        total_months = person.cal_months + person.acad_months + person.sum_months
        if data.project_info.sponsor_type == SponsorType.FEDERAL_OTHER and total_months > 2:
            warnings.append(
                f"NSF 2-month rule: {person.first_name} {person.last_name} has {total_months} months. "
                f"Verify program officer approval."
            )
        
        # Check NIH salary cap
        if data.project_info.sponsor_type == SponsorType.FEDERAL_NIH:
            cap = 228000 if person.cal_months > 0 else 161500
            if person.base_salary > cap:
                warnings.append(
                    f"NIH salary cap: {person.first_name} {person.last_name} salary ${person.base_salary:,.0f} "
                    f"exceeds cap ${cap:,.0f}. Ensure capped salary entered and cost share documented."
                )
    
    # === Grad Fringe Type ===
    if data.grad_fringe_type:
        safe_write('I43', data.grad_fringe_type.value)
    
    # === Other Personnel ===
    for i, person in enumerate(data.other_personnel):
        row = 53 + i  # Rows 53-72
        
        if person.prefix:
            safe_write(f'B{row}', person.prefix)
        safe_write(f'C{row}', person.first_name)
        if person.middle:
            safe_write(f'D{row}', person.middle)
        safe_write(f'E{row}', person.last_name)
        if person.suffix:
            safe_write(f'F{row}', person.suffix)
        safe_write(f'G{row}', person.role.value)
        
        if person.cal_months > 0:
            safe_write(f'J{row}', person.cal_months)
        if person.acad_months > 0:
            safe_write(f'K{row}', person.acad_months)
        if person.sum_months > 0:
            safe_write(f'L{row}', person.sum_months)
        
        safe_write(f'M{row}', person.requested_salary)
        
        # Generate and write justification
        justification = person.generate_justification()
        if justification:
            safe_write(f'P{row}', justification)
    
    # Save workbook
    wb.save(output_path)
    
    # Restore original external link files from template (openpyxl corrupts them)
    restored_count = preserve_external_links(template_path, output_path)
    
    return {
        'success': True,
        'validation_errors': [],
        'cells_written': cells_written,
        'warnings': warnings,
        'external_links_restored': restored_count
    }


def generate_verification_table(data: BudgetData) -> str:
    """Generate a markdown table showing what will be populated."""
    lines = ["## Budget Population Preview\n"]
    
    lines.append("### Project Information\n")
    lines.append("| Cell | Field | Value |")
    lines.append("|------|-------|-------|")
    lines.append(f"| I5 | Start Date | {data.project_info.start_date} |")
    lines.append(f"| K5 | Duration | {data.project_info.duration_years} years |")
    lines.append(f"| O5 | Sponsor Type | {data.project_info.sponsor_type.value} |")
    
    if data.senior_personnel:
        lines.append("\n### Senior Personnel\n")
        lines.append("| Row | Name | Role | Designation | Base Salary | Cal Mo | Acad Mo | Sum Mo |")
        lines.append("|-----|------|------|-------------|-------------|--------|---------|--------|")
        for i, p in enumerate(data.senior_personnel):
            row = 7 + i if i < 8 else 20 + (i - 8)
            lines.append(
                f"| {row} | {p.first_name} {p.last_name} | {p.role.value} | "
                f"{p.designation.value[:20]}... | ${p.base_salary:,.0f} | "
                f"{p.cal_months or '-'} | {p.acad_months or '-'} | {p.sum_months or '-'} |"
            )
    
    if data.grad_fringe_type:
        lines.append(f"\n### Grad Fringe Type (I43): {data.grad_fringe_type.value}\n")
    
    if data.other_personnel:
        lines.append("\n### Other Personnel\n")
        lines.append("| Row | Name | Role | Cal Mo | Acad Mo | Sum Mo | Salary |")
        lines.append("|-----|------|------|--------|---------|--------|--------|")
        for i, p in enumerate(data.other_personnel):
            row = 53 + i
            lines.append(
                f"| {row} | {p.first_name} {p.last_name} | {p.role.value[:20]}... | "
                f"{p.cal_months or '-'} | {p.acad_months or '-'} | {p.sum_months or '-'} | "
                f"${p.requested_salary:,.0f} |"
            )
    
    return "\n".join(lines)


def apply_out_year_changes(workbook_path: str, state: BudgetState, template_path: str = None) -> Dict[str, Any]:
    """
    Apply out-year personnel changes to an existing workbook.
    
    This function handles:
    - Removing personnel (zeroing out effort/salary)
    - Modifying personnel (changing effort, FTE, etc.)
    - Adding new personnel in out-years
    - Writing justifications for ALL rows with salary (including escalated ones)
    
    Args:
        workbook_path: Path to the workbook to modify
        state: BudgetState containing personnel_changes
        template_path: Path to original template (to read escalation rate)
        
    Returns:
        Dict with cells_written, warnings, etc.
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(workbook_path)
    cells_written = []
    warnings = []
    
    # Get escalation rate from template (cell D5)
    escalation_rate = 0.03  # Default
    if template_path:
        try:
            template_wb = load_workbook(template_path, data_only=True)
            template_sheet = template_wb['Personnel Yr 1']
            if template_sheet['D5'].value:
                escalation_rate = float(template_sheet['D5'].value)
        except:
            pass
    escalation_pct = f"{escalation_rate * 100:.0f}%"
    
    # Track which rows have Year 1 personnel (to know what escalates)
    y1_senior_rows = set()
    y1_other_rows = set()
    
    for i, p in enumerate(state.year1_senior):
        row = 7 + i if i < 8 else 20 + (i - 8)
        y1_senior_rows.add(row)
    
    for i, p in enumerate(state.year1_other):
        row = 53 + i
        y1_other_rows.add(row)
    
    for year in range(2, 6):
        if year > state.project_info.duration_years:
            continue
            
        sheet_name = f'Personnel Yr {year}'
        if sheet_name not in wb.sheetnames:
            warnings.append(f"Sheet {sheet_name} not found")
            continue
            
        sheet = wb[sheet_name]
        year_changes = state.get_changes_for_year(year)
        
        # Track which rows have changes this year
        rows_with_changes = {c.row for c in year_changes}
        
        # First, apply explicit changes
        for change in year_changes:
            row = change.row
            cell_values = change.get_cells_for_year(year)
            
            # For "add" changes, also write name and role if this is the effective year
            if change.change_type == "add" and year == change.effective_year:
                # Write name
                if "C" not in cell_values and change.person_name:
                    name_parts = change.person_name.split()
                    if len(name_parts) >= 2:
                        cell_values["C"] = name_parts[0]
                        cell_values["E"] = name_parts[-1]
                    else:
                        cell_values["C"] = change.person_name
                        cell_values["E"] = change.person_name
            
            # Write each cell value
            for col, value in cell_values.items():
                cell = f"{col}{row}"
                sheet[cell] = value
                cells_written.append((sheet_name, cell, value, change.description))
            
            # Write justification if provided (for adds and modifies)
            if change.justification and change.change_type in ["add", "modify"]:
                cell = f"P{row}"
                sheet[cell] = change.justification
                cells_written.append((sheet_name, cell, change.justification, "Justification"))
            
            # Log the change
            if change.change_type == "remove":
                warnings.append(f"Year {year}: Zeroed out {change.person_name} (Row {row}) - {change.description}")
            elif change.change_type == "modify":
                warnings.append(f"Year {year}: Modified {change.person_name} (Row {row}) - {change.description}")
            elif change.change_type == "add":
                warnings.append(f"Year {year}: Added {change.person_name} (Row {row}) - {change.description}")
        
        # Now, write justifications for rows that just escalate (no changes)
        # Senior personnel rows
        for row in y1_senior_rows:
            if row not in rows_with_changes:
                # Check if this row was removed in a prior year
                was_removed = False
                for c in state.personnel_changes:
                    if c.row == row and c.section == "senior" and c.change_type == "remove":
                        if c.applies_to_year(year):
                            was_removed = True
                            break
                
                if not was_removed:
                    cell = f"P{row}"
                    justification = f"Effort same as prior year. Salary escalated {escalation_pct} per cell D5."
                    sheet[cell] = justification
                    cells_written.append((sheet_name, cell, justification, "Escalation justification"))
        
        # Other personnel rows
        for row in y1_other_rows:
            if row not in rows_with_changes:
                # Check if this row was removed in a prior year
                was_removed = False
                for c in state.personnel_changes:
                    if c.row == row and c.section == "other" and c.change_type == "remove":
                        if c.applies_to_year(year):
                            was_removed = True
                            break
                
                if not was_removed:
                    cell = f"P{row}"
                    justification = f"Effort same as prior year. Salary escalated {escalation_pct} per cell D5."
                    sheet[cell] = justification
                    cells_written.append((sheet_name, cell, justification, "Escalation justification"))
        
        # Also handle rows added in previous years that continue
        for c in state.personnel_changes:
            if c.change_type == "add" and c.effective_year < year:
                if c.applies_to_year(year) and c.row not in rows_with_changes:
                    # This is a person added in a prior year, still active, no changes this year
                    cell = f"P{c.row}"
                    justification = f"Effort same as prior year. Salary escalated {escalation_pct} per cell D5."
                    sheet[cell] = justification
                    cells_written.append((sheet_name, cell, justification, "Escalation justification"))
    
    wb.save(workbook_path)
    
    return {
        'success': True,
        'cells_written': cells_written,
        'warnings': warnings
    }


def populate_budget_full(template_path: str, output_path: str, state: BudgetState) -> Dict[str, Any]:
    """
    Populate a complete budget from BudgetState, including out-year changes.
    
    This is the main entry point for full budget generation.
    
    Args:
        template_path: Path to the original OSP Budget Template
        output_path: Path for the output file
        state: Complete BudgetState with Year 1 data and out-year changes
        
    Returns:
        Dict with success status, cells written, warnings, etc.
    """
    # First, populate Year 1 using the existing function
    year1_data = state.to_budget_data()
    result = populate_budget(template_path, output_path, year1_data)
    
    if not result['success']:
        return result
    
    # Then apply out-year changes
    if state.personnel_changes or state.year1_senior or state.year1_other:
        out_year_result = apply_out_year_changes(output_path, state, template_path)
        
        # Merge results
        result['out_year_cells_written'] = out_year_result['cells_written']
        result['out_year_warnings'] = out_year_result['warnings']
        
        # Re-apply external links fix (we modified the file again)
        restored = preserve_external_links(template_path, output_path)
        result['external_links_restored'] = restored
    
    # Add change summary
    result['change_summary'] = state.summarize_changes()
    
    return result


# === MODEL CHECK ===

MODEL_WARNING = """
⚠️ MODEL REQUIREMENT WARNING ⚠️

This skill requires Claude Opus 4.5 for reliable budget template population.

The OSP Budget Template population involves:
- Complex natural language interpretation ("buy out one course" → 1.275 months)
- Precise dropdown value matching (exact strings required)
- Policy calculations (NIH caps, NSF 2-month rule, fringe rates)
- Different month bases for different personnel types
- Bespoke code generation for each unique budget

Using Haiku or Sonnet may result in:
- Incorrect cell values
- Formula corruption
- Policy violations
- Budget calculation errors

Please switch to Claude Opus 4.5 before proceeding.
"""

def check_model_recommendation() -> str:
    """Return a warning message about model requirements."""
    return MODEL_WARNING


# === EXAMPLE USAGE ===

if __name__ == "__main__":
    # Example: Create the budget we walked through
    data = BudgetData(
        project_info=ProjectInfo(
            start_date="7/1/26",
            duration_years=3,
            sponsor_type=SponsorType.FEDERAL_OTHER
        ),
        senior_personnel=[
            SeniorPerson(
                first_name="Jane",
                last_name="Smith",
                role=SeniorRole.PI,
                designation=Designation.TENURE_TRACK_ACAD_SUM,
                base_salary=105000,
                acad_months=1.275,
                sum_months=0.725
            ),
            SeniorPerson(
                first_name="Robert",
                last_name="Chen",
                role=SeniorRole.CO_PI,
                designation=Designation.NON_TENURE_FULL_TIME,
                base_salary=82000,
                cal_months=12
            ),
        ],
        grad_fringe_type=SponsorType.FEDERAL_OTHER,
        other_personnel=[
            OtherPerson(
                first_name="TBD",
                last_name="Postdoc",
                role=OtherRole.POSTDOC,
                cal_months=6,
                requested_salary=33000
            ),
            OtherPerson(
                first_name="TBD",
                last_name="Graduate Student 1",
                role=OtherRole.GRAD_ASSISTANT,
                acad_months=9,
                sum_months=2,
                requested_salary=35139
            ),
            OtherPerson(
                first_name="TBD",
                last_name="Graduate Student 2",
                role=OtherRole.GRAD_ASSISTANT,
                acad_months=4.5,
                requested_salary=14375
            ),
        ]
    )
    
    # Show preview
    print(generate_verification_table(data))
    
    # Validate
    errors = data.validate()
    if errors:
        print("\nValidation Errors:")
        for err in errors:
            print(f"  - {err}")
    else:
        print("\n✓ Data validated successfully")
